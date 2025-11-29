import io

import pytest
from openpyxl import load_workbook

from src.services.database.referrals.reports import generate_referral_report_excel



@pytest.mark.asyncio
async def test_generate_referral_report_excel(
        create_new_user,
        create_referral,
        create_income_from_referral
):
    # Создаём пользователя
    user = await create_new_user()

    # Создаём 3 рефералов
    ref_1, _, _ = await create_referral(owner_id=user.user_id)
    ref_2, _, _ = await create_referral(owner_id=user.user_id)
    ref_3, _, _ = await create_referral(owner_id=user.user_id)

    # Создаём начисления для ref_1
    income_1, _, _ = await create_income_from_referral(referral_user_id=ref_1.referral_id, owner_id=user.user_id)
    income_2, _, _ = await create_income_from_referral(referral_user_id=ref_1.referral_id, owner_id=user.user_id)
    income_3, _, _ = await create_income_from_referral(referral_user_id=ref_1.referral_id, owner_id=user.user_id)

    # Формируем Excel (байты)
    excel_bytes = await generate_referral_report_excel(user.user_id, 'ru')

    # Загружаем Excel из байтов через BytesIO
    wb = load_workbook(filename=io.BytesIO(excel_bytes))
    assert "referrals_report" in wb.sheetnames, "Лист referrals_report отсутствует"
    sheet = wb["referrals_report"]

    # Собираем все непустые значения в один список строк (строки -> значения)
    all_values = [str(cell)
                  for row in sheet.iter_rows(values_only=True)
                  for cell in row
                  if cell is not None]

    # Проверка: служебный блок содержит ID владельца
    assert any(str(user.user_id) == v for v in all_values), "ID владельца не найден в файле"

    # Проверяем, что в файле есть ID всех трёх рефералов
    assert any(str(ref_1.referral_id) == v for v in all_values), "ref_1 отсутствует в Excel"
    assert any(str(ref_2.referral_id) == v for v in all_values), "ref_2 отсутствует в Excel"
    assert any(str(ref_3.referral_id) == v for v in all_values), "ref_3 отсутствует в Excel"

    # Проверяем, что есть доходы (income_from_referral_id)
    assert any(str(income_1.income_from_referral_id) == v for v in all_values), "income_1 отсутствует"
    assert any(str(income_2.income_from_referral_id) == v for v in all_values), "income_2 отсутствует"
    assert any(str(income_3.income_from_referral_id) == v for v in all_values), "income_3 отсутствует"

    wb.close()